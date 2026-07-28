"""
count_segments_per_sentence.py — Segment-per-sentence statistics using spaCy
------------------------------------------------------------------------------

Input
- B I tagged .tsv files in Data/Segmentation_BI, inside *_DONE folders
- Language variety is derived from the name of the top-level folder:
    DE* → German, E* → English. Files in other folders are skipped with a warning.

Output
- .xlsx files in Data/:
    seg_per_sentence_E_HIST.xlsx
    seg_per_sentence_E_CURRENT.xlsx
    seg_per_sentence_DE_HIST.xlsx
    seg_per_sentence_DE_CURRENT.xlsx

    
Each workbook contains three sheets:
    Raw          — one row per (document, annotator, sentence)
    By Document  — aggregated stats per (document, annotator)
    Overall      — aggregated stats per annotator across all documents

Required spaCy models (install once):
    python -m spacy download en_core_web_sm
    python -m spacy download de_core_news_sm

Usage:
    python count_segments_per_sentence.py [--root <path>] [--dry-run]

Options:
    --root     Directory containing the Data/ folder (default: cwd).
    --dry-run  Discover files and print them without running analysis.
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import spacy
from spacy.tokens import Doc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# spaCy model registry  (language code → model name)
# ---------------------------------------------------------------------------

SPACY_MODELS = {
    "EN": "en_core_web_sm",
    "DE": "de_core_news_sm",
}

_nlp_cache: dict[str, spacy.Language] = {}


def get_nlp(lang: str) -> spacy.Language:
    """
    Return a minimal spaCy pipeline for *lang* that only detects sentence
    boundaries.

    We use the rule-based 'sentencizer' rather than the neural 'senter'
    because:
      - 'senter' requires 'tok2vec' to run first; disabling tok2vec (for
        speed) leaves sentence boundaries unset and raises E030.
      - Our input is already tokenised, so a punctuation-based split is
        both correct and fast.

    The full model is still loaded so that the vocab (and any custom
    tokeniser settings) remain language-appropriate.
    """
    if lang in _nlp_cache:
        return _nlp_cache[lang]

    model_name = SPACY_MODELS.get(lang)
    if model_name is None:
        raise ValueError(
            f"No spaCy model configured for language '{lang}'. "
            f"Add an entry to SPACY_MODELS in the script."
        )

    try:
        nlp = spacy.load(model_name)
    except OSError:
        sys.exit(
            f"\nspaCy model '{model_name}' not found.\n"
            f"Install it with:\n"
            f"    python -m spacy download {model_name}\n"
        )

    # Disable every component — no tagging, parsing, or NER
    if nlp.pipe_names:
        nlp.disable_pipes(*nlp.pipe_names)

    # Enable or add the rule-based sentencizer (splits on sentence-ending punctuation)
    if "sentencizer" in nlp.pipe_names:
        nlp.enable_pipe("sentencizer")
    else:
        nlp.add_pipe("sentencizer")

    _nlp_cache[lang] = nlp
    return nlp

# Folder-name helpers  (mirrors batch_evaluate_BI.py)
# ---------------------------------------------------------------------------

def annotator_from_folder(name: str) -> str:
    return name.removesuffix("_DONE")


def language_from_top_folder(top: str) -> str:
    """
    Derive the two-letter language code from the variety folder name.

    Folder names start with a single letter indicating language:
        E…  → EN  (English)
        D…  → DE  (German)
    Any other prefix is a configuration error
    """
    u = top.upper()
    if u.startswith("E"):
        return "EN"
    if u.startswith("D"):
        return "DE"
    raise ValueError(
        f"Cannot determine language from folder name '{top}'. "
        f"Expected a name starting with 'E' (English) or 'D' (German)."
    )


def variety_from_top_folder(top: str) -> str:
    u = top.upper()
    lang = "DE" if u.startswith("DE") else "E"
    if "HIST" in u:
        return f"{lang}_HIST"
    if "CURR" in u:
        return f"{lang}_CURRENT"
    return lang


def doc_label(path: Path, segmentation_bi: Path) -> str:
    try:
        rel   = path.relative_to(segmentation_bi)
        parts = list(rel.parts)
        return "/".join(parts[: len(parts) - 2] + [parts[-1]])
    except ValueError:
        return path.name


# File discovery
# ---------------------------------------------------------------------------

def find_files(
    segmentation_bi: Path,
) -> list[tuple[Path, str, str, str]]:
    """
    Return every .tagged.tsv file inside a *_DONE folder as:
        (path, annotator, language, variety)
    """
    results: list[tuple[Path, str, str, str]] = []

    for tsv in sorted(segmentation_bi.rglob("*.tagged.tsv")):
        if not tsv.parent.name.endswith("_DONE"):
            continue
        try:
            rel = tsv.relative_to(segmentation_bi)
        except ValueError:
            continue
        top_folder = rel.parts[0]
        annotator  = annotator_from_folder(tsv.parent.name)
        language   = language_from_top_folder(top_folder)
        variety    = variety_from_top_folder(top_folder)
        results.append((tsv, annotator, language, variety))

    return results


# TSV parsing
# ---------------------------------------------------------------------------

def read_tsv(path: Path) -> list[tuple[str, str]]:
    """
    Read a .tagged.tsv file and return a list of (token, tag) pairs.
    Blank lines are skipped.
    """
    pairs: list[tuple[str, str]] = []
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            token = parts[0]
            tag   = parts[1].strip() if len(parts) > 1 else "O"
            pairs.append((token, tag))
    return pairs


# Segment detection  (same BIO logic as batch_evaluate_BI.py)
# ---------------------------------------------------------------------------

def tag_to_spans(token_tags: list[tuple[str, str]]) -> list[tuple[int, int]]:
    """
    Convert a BIO-tagged token list to a list of (start, end) span tuples
    (end is exclusive, matching Python slice conventions).
    """
    spans:   list[tuple[int, int]] = []
    start:   int | None            = None

    for i, (_, tag) in enumerate(token_tags):
        u = tag.upper()
        if u == "O" or u == "":
            if start is not None:
                spans.append((start, i))
                start = None
        elif u.startswith("I-") or u == "I":
            if start is None:
                start = i   # recover from missing B
        else:                # B- or any non-O/non-I tag
            if start is not None:
                spans.append((start, i))
            start = i

    if start is not None:
        spans.append((start, len(token_tags)))

    return spans


# Core analysis: sentences × segments
# ---------------------------------------------------------------------------

def analyse_file(
    path: Path,
    annotator: str,
    language: str,
    variety: str,
    segmentation_bi: Path,
) -> list[dict]:
    """
    For one .tagged.tsv file, use spaCy to split into sentences and
    count how many annotated segments (wholly or partially) overlap each
    sentence.  Returns a list of row dicts (one per sentence).

    Overlap rule: a segment is counted for a sentence if its token range
    intersects the sentence's token range — i.e. it starts or ends inside
    the sentence, or completely spans it.
    """
    token_tags = read_tsv(path)
    if not token_tags:
        return []

    tokens = [tok for tok, _ in token_tags]
    spans  = tag_to_spans(token_tags)

    # Build a spaCy Doc from the pre-tokenised list, then detect sentences
    nlp = get_nlp(language)
    doc = Doc(nlp.vocab, words=tokens)
    # Apply only the sentence boundary detector
    for _, component in nlp.pipeline:
        doc = component(doc)

    doc_id = doc_label(path, segmentation_bi)
    rows: list[dict] = []

    for sent_idx, sent in enumerate(doc.sents):
        sent_start = sent.start
        sent_end   = sent.end   # exclusive

        # Count segments that overlap this sentence
        seg_count = sum(
            1 for (s_start, s_end) in spans
            if s_start < sent_end and s_end > sent_start
        )

        rows.append({
            "document":        doc_id,
            "variety":         variety,
            "annotator":       annotator,
            "sentence_idx":    sent_idx + 1,             # 1-based for readability
            "sentence_text":   " ".join(tokens[sent_start:sent_end]),
            "sent_len_tokens": sent_end - sent_start,
            "segment_count":   seg_count,
        })

    return rows


# ---------------------------------------------------------------------------
# Aggregation helpers
# ---------------------------------------------------------------------------

def _safe_mean(s: pd.Series) -> str:
    return f"{s.mean():.2f}" if not s.empty else "—"


def _safe_median(s: pd.Series) -> str:
    return f"{np.median(s.values):.2f}" if not s.empty else "—"


def _safe_max(s: pd.Series) -> str:
    return str(int(s.max())) if not s.empty else "—"


def _pct_sents(grp: pd.DataFrame, n: int) -> str:
    """Percentage of sentences in *grp* where segment_count == *n*."""
    if grp.empty:
        return "—"
    pct = (grp["segment_count"] == n).mean() * 100
    return f"{pct:.1f}%"


def _pct_sents_ge(grp: pd.DataFrame, n: int) -> str:
    """Percentage of sentences where segment_count >= *n*."""
    if grp.empty:
        return "—"
    pct = (grp["segment_count"] >= n).mean() * 100
    return f"{pct:.1f}%"


def build_by_document(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate raw sentence rows to one row per (document, annotator)."""
    rows: list[dict] = []
    for (doc, ann), grp in df.groupby(["document", "annotator"], sort=True):
        sc = grp["segment_count"]
        rows.append({
            "document":            doc,
            "variety":             grp["variety"].iloc[0],
            "annotator":           ann,
            "total_sentences":     len(grp),
            "total_segments":      int(sc.sum()),
            "mean_segs_per_sent":  round(sc.mean(), 3),
            "median_segs_per_sent":round(float(np.median(sc.values)), 3),
            "max_segs_per_sent":   int(sc.max()),
            "std_segs_per_sent":   round(float(sc.std(ddof=1)) if len(sc) > 1 else 0.0, 3),
            "pct_sents_0_segs":    f"{(sc == 0).mean() * 100:.1f}%",
            "pct_sents_1_seg":     f"{(sc == 1).mean() * 100:.1f}%",
            "pct_sents_2plus_segs":f"{(sc >= 2).mean() * 100:.1f}%",
        })
    return pd.DataFrame(rows)


def build_overall(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate raw sentence rows to one row per annotator."""
    rows: list[dict] = []
    for ann, grp in df.groupby("annotator", sort=True):
        sc = grp["segment_count"]
        rows.append({
            "annotator":           ann,
            "total_documents":     grp["document"].nunique(),
            "total_sentences":     len(grp),
            "total_segments":      int(sc.sum()),
            "mean_segs_per_sent":  round(sc.mean(), 3),
            "median_segs_per_sent":round(float(np.median(sc.values)), 3),
            "max_segs_per_sent":   int(sc.max()),
            "std_segs_per_sent":   round(float(sc.std(ddof=1)) if len(sc) > 1 else 0.0, 3),
            "pct_sents_0_segs":    f"{(sc == 0).mean() * 100:.1f}%",
            "pct_sents_1_seg":     f"{(sc == 1).mean() * 100:.1f}%",
            "pct_sents_2plus_segs":f"{(sc >= 2).mean() * 100:.1f}%",
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# XLSX styling
# ---------------------------------------------------------------------------

_HDR_FILL    = PatternFill("solid", start_color="1F4E79")
_ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
_SUBHDR_FILL = PatternFill("solid", start_color="2E75B6")
_AGG_FILL    = PatternFill("solid", start_color="E2EFDA")   # light green for agg sheets
_AGG_HDR     = PatternFill("solid", start_color="375623")   # dark green
_HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT   = Font(name="Arial", size=10)
_BOLD_FONT   = Font(name="Arial", bold=True, size=10)
_THIN_SIDE   = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)


def _style_header(ws, row: int, n_cols: int,
                  fill: PatternFill = None) -> None:
    fill = fill or _HDR_FILL
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font      = _HDR_FONT
        c.fill      = fill
        c.border    = _THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)


def _style_body(cell, alt: bool = False,
                alt_fill: PatternFill = None) -> None:
    cell.font      = _BODY_FONT
    cell.border    = _THIN_BORDER
    cell.fill      = (alt_fill or _ALT_FILL) if alt else PatternFill()
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_w: int = 8, max_w: int = 50) -> None:
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max(length + 2, min_w), max_w)


def _write_table(ws, df: pd.DataFrame,
                 hdr_fill: PatternFill = None,
                 alt_fill: PatternFill = None,
                 freeze_col: int = 1) -> None:
    """Write *df* as a styled table into *ws* (1-indexed, starting at row 1)."""
    headers = list(df.columns)
    ws.append(headers)
    _style_header(ws, 1, len(headers), fill=hdr_fill)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = ws.cell(row=2, column=freeze_col).coordinate

    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        alt = (i % 2 == 0)
        for col in range(1, len(headers) + 1):
            _style_body(ws.cell(row=i, column=col),
                        alt=alt, alt_fill=alt_fill)

    _autofit(ws)


# Per-variety XLSX writer
# ---------------------------------------------------------------------------
def write_combined_xlsx(path: Path, df_raw: pd.DataFrame) -> None:
    """
    Write a single three-sheet workbook covering all varieties.
    The variety column is present in every sheet for filtering.
    """
    df_by_doc  = build_by_document(df_raw)
    df_overall = build_overall(df_raw)

    raw_display = df_raw.rename(columns={
        "sentence_idx":    "sent_#",
        "sent_len_tokens": "sent_len (tokens)",
        "segment_count":   "segments_in_sent",
    })

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name in ("Raw", "By Document", "Overall"):
            pd.DataFrame().to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(path)

    ws = wb["Raw"]
    ws.delete_rows(1, ws.max_row)
    _write_table(ws, raw_display, hdr_fill=_HDR_FILL, alt_fill=_ALT_FILL)
    text_col = raw_display.columns.get_loc("sentence_text") + 1
    for row in ws.iter_rows(min_row=2, min_col=text_col, max_col=text_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=False)

    ws = wb["By Document"]
    ws.delete_rows(1, ws.max_row)
    _write_table(ws, df_by_doc, hdr_fill=_SUBHDR_FILL, alt_fill=_ALT_FILL)

    ws = wb["Overall"]
    ws.delete_rows(1, ws.max_row)
    _write_table(ws, df_overall, hdr_fill=_AGG_HDR, alt_fill=_AGG_FILL)
    ws.insert_rows(1)
    varieties = ", ".join(sorted(df_raw["variety"].unique()))
    tc = ws.cell(row=1, column=1,
                 value=f"Segments per Sentence — {varieties}")
    tc.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc.fill      = _AGG_HDR
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1,   end_column=len(df_overall.columns),
    )
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A3"

    wb.save(path)


# Terminal summary
# ---------------------------------------------------------------------------

def print_variety_summary(variety: str, df: pd.DataFrame) -> None:
    sc = df["segment_count"]
    print(f"\n  {variety}  ({df['document'].nunique()} docs, "
          f"{df['annotator'].nunique()} annotator(s), "
          f"{len(df)} sentences)")
    print(f"    Mean segments/sent : {sc.mean():.2f}")
    print(f"    Median             : {np.median(sc.values):.2f}")
    print(f"    Max                : {int(sc.max())}")
    print(f"    Sents with 0 segs  : {(sc == 0).mean() * 100:.1f}%")
    print(f"    Sents with 1 seg   : {(sc == 1).mean() * 100:.1f}%")
    print(f"    Sents with 2+ segs : {(sc >= 2).mean() * 100:.1f}%")
    for ann, grp in df.groupby("annotator"):
        s = grp["segment_count"]
        print(f"    [{ann}]  mean={s.mean():.2f}  "
              f"median={np.median(s.values):.2f}  "
              f"docs={grp['document'].nunique()}")



# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Count segments per spaCy-detected sentence in BI TSV files"
    )
    parser.add_argument(
        "--root", type=Path, default=Path.cwd(),
        help="Directory containing the Data/ folder (default: cwd)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Discover files and print them without running analysis",
    )
    args = parser.parse_args()

    segmentation_bi = args.root / "Data" / "Segmentation_BI"
    out_dir         = args.root / "Data"

    if not segmentation_bi.is_dir():
        sys.exit(f"Error: directory not found: {segmentation_bi}")

    print(f"Scanning {segmentation_bi}…")
    files = find_files(segmentation_bi)

    if not files:
        print("No .tagged.tsv files found.")
        return

    print(f"Found {len(files)} file(s).\n")

    if args.dry_run:
        for path, ann, lang, variety in files:
            print(f"  [{lang}/{variety}]  {ann}:  {path.name}")
        print("\nDry run — nothing analysed.")
        return

    variety_rows: dict[str, list[dict]] = defaultdict(list)

    for i, (path, ann, lang, variety) in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {ann} / {variety}: {path.name}")
        try:
            rows = analyse_file(path, ann, lang, variety, segmentation_bi)
        except Exception as exc:
            print(f"  Error: {exc}")
            continue

        total_segs = sum(r["segment_count"] for r in rows)
        print(f"  {len(rows)} sentences, {total_segs} segment(s) found")
        variety_rows[variety].extend(rows)

    if not variety_rows:
        print("\nNo data produced.")
        return

    all_rows = [row for rows in variety_rows.values() for row in rows]
    df_all   = pd.DataFrame(all_rows)
    out_path = out_dir / "seg_per_sentence.xlsx"

    sep = "=" * 60
    print(f"\n{sep}")
    print(f"  SEGMENT-PER-SENTENCE SUMMARY")
    print(f"{sep}")

    for variety in sorted(variety_rows):
        df = pd.DataFrame(variety_rows[variety])
        print_variety_summary(variety, df)

    write_combined_xlsx(out_path, df_all)
    print(f"\nWritten → {out_path}")
    print(f"\n{sep}\n")


if __name__ == "__main__":
    main()