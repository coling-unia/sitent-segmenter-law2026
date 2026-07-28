"""
batch_evaluate_BI.py — Batch Inter-Annotator Agreement Evaluation
---------------------------------------------------------------

Input
  - Recursively scans Data/Segmentation_BI
  - Includes files ending in:
        *.tagged.tsv (default expected format)
        *.tsv        (accepted for backward compatibility)
  - By default, only folders ending with "_DONE" are treated as annotator folders
    (configurable via --suffix)
  - Groups files that represent the same document (same path modulo the
annotator folder)


Output

  Data/evaluation_results.csv   — one row per document pair, all counts
                                   also expressed as % of the span union
  Data/evaluation_results.xlsx  — same data plus a Summary sheet and a
                                   By Variety sheet (E_HIST / E_CURRENT /
                                   DE_HIST / DE_CURRENT)
  Data/non_matches.csv          — one row per non-matching span (partial
                                   overlaps and fully unmatched spans)


Usage:
    python batch_evaluate_BI.py [OPTIONS]

Options:
    --root <path>              Directory containing the Data/ folder
                               (default: current working directory)

    --dry-run                  Discover and print annotator pairs without
                               running evaluation

    --suffix <str>             Annotator folder suffix (default: _DONE)

    --include-dir <str ...>    Only include files whose path contains any
                               of these substrings (e.g. DE E_CURRENT)

    --exclude-dir <str ...>    Exclude files whose path contains any of
                               these substrings (e.g. WIP OLD)

    --folder-whitelist <str ...>
                               Only process annotator folders containing
                               these substrings (e.g. DONE FINAL)

    --folder-blacklist <str ...>
                               Skip annotator folders containing these
                               substrings (e.g. WIP TEMP, model if only 
                               human IAA is of interest)

"""

import argparse
import sys
from collections import defaultdict
from itertools import combinations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from evaluate_unitizing import evaluate_pair, EvaluationResult
import time
timestr = time.strftime("%Y%m%d-%H%M%S")


# File discovery + pairing
# ---------------------------------------------------------------------------
def match_any(name: str, patterns: list[str] | None) -> bool:
    if not patterns:
        return True
    name_upper = name.upper()
    return any(p.upper() in name_upper for p in patterns)


def match_none(name: str, patterns: list[str] | None) -> bool:
    if not patterns:
        return True
    name_upper = name.upper()
    return not any(p.upper() in name_upper for p in patterns)


def annotator_from_folder(folder_name: str, suffix: str) -> str:
    return folder_name.removesuffix(suffix)


def language_from_top_folder(top_folder: str) -> str:
    upper = top_folder.upper()
    if upper.startswith("DE"):
        return "DE"
    if upper.startswith("EN"):
        return "EN"
    return top_folder


def variety_from_top_folder(top_folder: str) -> str:
    """
    Map the raw top-level folder name to one of the four canonical varieties:
    E_HIST, E_CURRENT, DE_HIST, DE_CURRENT.

    Expected folder naming conventions (case-insensitive):
      DE_HIST / DE_HISTORICAL  → DE_HIST
      DE_CURRENT / DE_CURR     → DE_CURRENT
      EN_HIST / E_HIST …       → E_HIST
      EN_CURRENT / E_CURRENT … → E_CURRENT
    Falls back to the raw folder name if nothing matches.
    """
    upper = top_folder.upper()
    is_de = upper.startswith("DE")
    lang  = "DE" if is_de else "E"
    if "HIST" in upper:
        return f"{lang}_HIST"
    if "CURR" in upper:
        return f"{lang}_CURRENT"
    return lang


def find_pairs(
    segmentation_bi: Path,
    suffix: str = "_DONE",
    include_dir: list[str] | None = None,
    exclude_dir: list[str] | None = None,
    folder_whitelist: list[str] | None = None,
    folder_blacklist: list[str] | None = None,
) -> list[tuple[Path, Path, str, str, str, str]]:
    

    groups: dict[str, list[tuple[str, Path]]] = defaultdict(list)

    for tsv in sorted(segmentation_bi.rglob("*.tagged.tsv")):
        folder_name = tsv.parent.name

        # --- suffix filter (replaces hardcoded _DONE) ---
        if suffix and not folder_name.endswith(suffix):
            continue

        # --- folder whitelist/blacklist ---
        if not match_any(folder_name, folder_whitelist):
            continue
        if not match_none(folder_name, folder_blacklist):
            continue

        # --- path-level include/exclude ---
        full_path_str = str(tsv)
        if not match_any(full_path_str, include_dir):
            continue
        if not match_none(full_path_str, exclude_dir):
            continue

        annotator = annotator_from_folder(folder_name, suffix)
        try:
            rel = tsv.relative_to(segmentation_bi)
        except ValueError:
            continue
        parts = list(rel.parts)
        done_idx = next(
            (i for i, p in enumerate(parts) if p.endswith(suffix)),
            None
        )
        if done_idx is None:
            continue
        key_parts = parts[:done_idx] + ["__DOC__"] + parts[done_idx + 1:]
        doc_key = "/".join(parts[:done_idx] + [tsv.name])
        groups[doc_key].append((annotator, tsv))

    pairs: list[tuple[Path, Path, str, str, str, str]] = []
    for doc_key, entries in sorted(groups.items()):
        if len(entries) < 2:
            continue
        top_folder = doc_key.split("/")[0]
        lang    = language_from_top_folder(top_folder)
        variety = variety_from_top_folder(top_folder)
        for (ann_a, path_a), (ann_b, path_b) in combinations(entries, 2):
            pairs.append((path_a, path_b, ann_a, ann_b, lang, variety))

    return pairs


def find_text_mismatch(path_a: Path, path_b: Path) -> str:
    """
    Read the token column (col 0) from both .tagged.tsv files and return
    a human-readable description of the first divergence, e.g.:
        token 42: 'Hund' (A) vs 'Katze' (B)  |  A has 120 tokens, B has 118 tokens
    Falls back to a plain length diff if the files cannot be parsed.
    """
    def load_tokens(p: Path) -> list[str]:
        tokens: list[str] = []
        with p.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.rstrip("\n")
                if not line:
                    continue
                tokens.append(line.split("\t")[0])
        return tokens

    try:
        toks_a = load_tokens(path_a)
        toks_b = load_tokens(path_b)
    except Exception as exc:
        return f"(could not read files for diff: {exc})"

    len_a, len_b = len(toks_a), len(toks_b)
    min_len = min(len_a, len_b)

    first_diff: str | None = None
    for i in range(min_len):
        if toks_a[i] != toks_b[i]:
            first_diff = (
                f"first divergence at token {i + 1}: "
                f"\'{toks_a[i]}\' (A) vs \'{toks_b[i]}\' (B)"
            )
            break

    length_info = (
        f"A has {len_a} tokens, B has {len_b} tokens"
        if len_a != len_b
        else f"both have {len_a} tokens"
    )

    if first_diff:
        return f"{first_diff}  |  {length_info}"
    if len_a != len_b:
        extra_side  = "A" if len_a > len_b else "B"
        extra_token = toks_a[min_len] if len_a > len_b else toks_b[min_len]
        return (
            f"texts agree for first {min_len} tokens, "
            f"then {extra_side} has extra token \'{extra_token}\'  |  {length_info}"
        )
    return "(mismatch reported by evaluator but no token difference found)"


def doc_label(path: Path, segmentation_bi: Path) -> str:
    """Clean document label — path relative to Segmentation_BI without the annotator folder."""
    try:
        rel = path.relative_to(segmentation_bi)
        parts = list(rel.parts)
        return "/".join(parts[:len(parts) - 2] + [parts[-1]])
    except ValueError:
        return path.name


# Segment-length statistics
# ---------------------------------------------------------------------------

def read_segment_lengths(path: Path) -> list[int]:
    """
    Parse a .tagged.tsv file (token<TAB>tag per line, blank lines ignored)
    and return the token-length of every annotated segment.

    Supports BIO tags (B-*, I-*), plain B/I, and IOB2.  Any tag that is not
    a continuation tag (i.e. not starting with "I") is treated as a potential
    segment start: it starts a new segment only when the tag is non-O.
    """
    lengths: list[int] = []
    current_len = 0

    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.rstrip("\n")
            if not line:
                # Blank lines can serve as sentence boundaries in some formats;
                # flush any open segment.
                if current_len > 0:
                    lengths.append(current_len)
                    current_len = 0
                continue

            parts = line.split("\t")
            tag = parts[1].strip() if len(parts) > 1 else "O"
            tag_upper = tag.upper()

            if tag_upper == "O" or tag_upper == "":
                # Outside — close any open segment
                if current_len > 0:
                    lengths.append(current_len)
                    current_len = 0
            elif tag_upper.startswith("I-") or tag_upper == "I":
                # Inside — continue current segment (or start one if malformed)
                if current_len == 0:
                    current_len = 1   # recover from missing B tag
                else:
                    current_len += 1
            else:
                # B- or any non-O/non-I tag → start of a new segment
                if current_len > 0:
                    lengths.append(current_len)
                current_len = 1

    if current_len > 0:
        lengths.append(current_len)

    return lengths


def segment_length_stats(lengths: list[int], suffix: str) -> dict:
    """
    Return a dict of descriptive statistics for *lengths*, with keys
    suffixed by *suffix* (e.g. '_a' or '_b').

    Keys produced:
        seg_n_<suffix>      — number of segments (sanity check vs spans_*)
        seg_mean_<suffix>   — mean tokens per segment
        seg_median_<suffix> — median tokens per segment
        seg_min_<suffix>    — shortest segment (tokens)
        seg_max_<suffix>    — longest segment (tokens)
        seg_std_<suffix>    — std-dev of segment lengths
    """
    if not lengths:
        return {
            f"seg_n{suffix}":      0,
            f"seg_mean{suffix}":   None,
            f"seg_median{suffix}": None,
            f"seg_min{suffix}":    None,
            f"seg_max{suffix}":    None,
            f"seg_std{suffix}":    None,
        }
    arr = np.array(lengths, dtype=float)
    return {
        f"seg_n{suffix}":      len(lengths),
        f"seg_mean{suffix}":   round(float(arr.mean()), 2),
        f"seg_median{suffix}": round(float(np.median(arr)), 2),
        f"seg_min{suffix}":    int(arr.min()),
        f"seg_max{suffix}":    int(arr.max()),
        f"seg_std{suffix}":    round(float(arr.std(ddof=1)) if len(lengths) > 1 else 0.0, 2),
    }


# Stats row  (evaluation_results.csv / Results sheet)
# ---------------------------------------------------------------------------

def result_to_stats_row(
    result: EvaluationResult,
    path_a: Path,
    path_b: Path,
    annotator_a: str,
    annotator_b: str,
    language: str,
    variety: str,
    segmentation_bi: Path,
) -> dict:
    union        = result.spans_a + result.spans_b - result.exact
    pct_of_union = lambda n: round(n / union * 100, 2) if union else 0.0
    pct_of_a     = lambda n: round(n / result.spans_a * 100, 2) if result.spans_a else 0.0
    pct_of_b     = lambda n: round(n / result.spans_b * 100, 2) if result.spans_b else 0.0

    only_a = len(result.no_overlap_a)
    only_b = len(result.no_overlap_b)

    # -- segment-length statistics (read directly from the source TSV files) --
    len_stats_a = segment_length_stats(read_segment_lengths(path_a), "_a")
    len_stats_b = segment_length_stats(read_segment_lengths(path_b), "_b")

    return {
        "document":              doc_label(path_a, segmentation_bi),
        "language":              language,
        "variety":               variety,
        "annotator_a":           annotator_a,
        "annotator_b":           annotator_b,
        # ---- agreement counts ----
        "spans_a":               result.spans_a,
        "spans_b":               result.spans_b,
        "union":                 union,
        "exact_matches":         result.exact,
        "exact_pct_a":           pct_of_a(result.exact),
        "exact_pct_b":           pct_of_b(result.exact),
        "partial_overlaps":      result.partial_overlaps,
        "partial_overlaps_pct":  pct_of_union(result.partial_overlaps),
        "only_a":                only_a,
        "only_a_pct":            pct_of_union(only_a),
        "only_b":                only_b,
        "only_b_pct":            pct_of_union(only_b),
        "precision":             round(result.precision, 4),
        "recall":                round(result.recall, 4),
        "f1":                    round(result.f1, 4),
        "ws_unified":            result.ws_unified,
        # ---- segment-length stats: annotator A ----
        **len_stats_a,
        # ---- segment-length stats: annotator B ----
        **len_stats_b,
    }


# Non-match rows  (non_matches.csv)
# ---------------------------------------------------------------------------

def result_to_nonmatch_rows(
    result: EvaluationResult,
    path_a: Path,
    annotator_a: str,
    annotator_b: str,
    language: str,
    variety: str,
    segmentation_bi: Path,
) -> list[dict]:
    rows: list[dict] = []
    doc = doc_label(path_a, segmentation_bi)

    def span_text(span: tuple[int, int]) -> str:
        return " ".join(tok for tok, _ in result.tagged[span[0]:span[1]])

    base = {
        "document":    doc,
        "language":    language,
        "variety":     variety,
        "annotator_a": annotator_a,
        "annotator_b": annotator_b,
    }

    for sa, sb in result.partial_pairs:
        rows.append({
            **base,
            "type":          "partial_overlap",
            "span_a_tokens": f"{sa[0]}:{sa[1]}",
            "span_b_tokens": f"{sb[0]}:{sb[1]}",
            "text_a":        span_text(sa),
            "text_b":        span_text(sb),
        })

    for span in result.no_overlap_a:
        rows.append({
            **base,
            "type":          "only_a",
            "span_a_tokens": f"{span[0]}:{span[1]}",
            "span_b_tokens": "",
            "text_a":        span_text(span),
            "text_b":        "",
        })

    for span in result.no_overlap_b:
        rows.append({
            **base,
            "type":          "only_b",
            "span_a_tokens": "",
            "span_b_tokens": f"{span[0]}:{span[1]}",
            "text_a":        "",
            "text_b":        span_text(span),
        })

    return rows


# Terminal summary
# ---------------------------------------------------------------------------

def _avg_exact(grp: pd.DataFrame) -> float:
    return grp[["exact_pct_a", "exact_pct_b"]].mean(axis=1).mean()


def _pooled_seg_mean(grp: pd.DataFrame, suffix: str) -> str:
    """Weighted mean of per-document segment-length means, weighted by segment count."""
    n_col    = f"seg_n{suffix}"
    mean_col = f"seg_mean{suffix}"
    valid = grp[[n_col, mean_col]].dropna()
    if valid.empty:
        return "—"
    weighted = (valid[n_col] * valid[mean_col]).sum() / valid[n_col].sum()
    return f"{weighted:.1f}"


def print_summary(df: pd.DataFrame, skipped: list[str]) -> None:
    sep = "=" * 60
    print(f"\n{sep}")
    print(f"  BATCH EVALUATION SUMMARY")
    print(f"{sep}")
    print(f"  Document pairs evaluated: {len(df)}")
    if skipped:
        print(f"  Pairs skipped (text mismatch): {len(skipped)}")
        for s in skipped:
            print(f"    • {s}")

    if df.empty:
        print("  No results to summarise.")
        print(f"{sep}\n")
        return

    print(f"\n  Overall (macro-averaged across all pairs):")
    print(f"    Precision : {df['precision'].mean():.3f}")
    print(f"    Recall    : {df['recall'].mean():.3f}")
    print(f"    F1        : {df['f1'].mean():.3f}")
    print(f"    Exact %   : {_avg_exact(df):.1f}%  (avg of A and B)")
    print(f"    Seg len A : mean={_pooled_seg_mean(df, '_a')} tokens")
    print(f"    Seg len B : mean={_pooled_seg_mean(df, '_b')} tokens")

    if df["language"].nunique() > 1:
        print(f"\n  By language:")
        for lang, grp in df.groupby("language"):
            print(f"    {lang}:  F1={grp['f1'].mean():.3f}  "
                  f"Exact%={_avg_exact(grp):.1f}%  "
                  f"n={len(grp)}")

    pair_col = df.apply(
        lambda r: " / ".join(sorted([r["annotator_a"], r["annotator_b"]])), axis=1
    )
    if pair_col.nunique() > 1:
        print(f"\n  By annotator pair:")
        for pair, grp in df.groupby(pair_col):
            print(f"    {pair}:  F1={grp['f1'].mean():.3f}  "
                  f"Exact%={_avg_exact(grp):.1f}%  "
                  f"n={len(grp)}")

    print(f"\n{sep}\n")


# XLSX export
# ---------------------------------------------------------------------------

# Shared palette
_HDR_FILL    = PatternFill("solid", start_color="1F4E79")   # dark blue
_ALT_FILL    = PatternFill("solid", start_color="D6E4F0")   # light blue
_SUBHDR_FILL = PatternFill("solid", start_color="2E75B6")   # mid blue
_SEG_FILL    = PatternFill("solid", start_color="E2EFDA")   # light green (seg-len columns)
_SEG_HDR_FILL= PatternFill("solid", start_color="375623")   # dark green (seg-len header)
_HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT   = Font(name="Arial", size=10)
_BOLD_FONT   = Font(name="Arial", bold=True, size=10)
_THIN_SIDE   = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE,  bottom=_THIN_SIDE)

# Columns that belong to the segment-length block (matched by prefix)
_SEG_LEN_PREFIX = "seg_"


def _is_seg_col(col_name: str) -> bool:
    return str(col_name).startswith(_SEG_LEN_PREFIX)


def _style_header_row(ws, row: int, col_names: list[str]) -> None:
    for col_idx, col_name in enumerate(col_names, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font   = _HDR_FONT
        cell.fill   = _SEG_HDR_FILL if _is_seg_col(col_name) else _HDR_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def _style_body_cell(cell, col_name: str = "", alt: bool = False) -> None:
    cell.font   = _BODY_FONT
    cell.border = _THIN_BORDER
    if _is_seg_col(col_name):
        # Segment-len columns always get their own subtle green tint
        cell.fill = _SEG_FILL
    else:
        cell.fill = _ALT_FILL if alt else PatternFill()
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_width: int = 8, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_width), max_width)


def _write_results_sheet(ws, df: pd.DataFrame) -> None:
    """Sheet 1 — raw results table, one row per document pair.
    """
    headers = list(df.columns)
    ws.append(headers)
    _style_header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


    seg_indices = [i + 1 for i, h in enumerate(headers) if _is_seg_col(h)]
    if seg_indices:
        ws.insert_rows(1)
        first_seg = seg_indices[0]
        last_seg  = seg_indices[-1]

        banner_cell = ws.cell(row=1, column=first_seg,
                              value="Tokens per Segment")
        banner_cell.font      = Font(name="Arial", bold=True,
                                     color="FFFFFF", size=10)
        banner_cell.fill      = _SEG_HDR_FILL
        banner_cell.alignment = Alignment(horizontal="center",
                                          vertical="center")
        banner_cell.border    = _THIN_BORDER
        if last_seg > first_seg:
            ws.merge_cells(
                start_row=1, start_column=first_seg,
                end_row=1,   end_column=last_seg
            )

        for col_idx in range(1, len(headers) + 1):
            if col_idx not in seg_indices:
                c = ws.cell(row=1, column=col_idx)
                c.fill   = _HDR_FILL
                c.border = _THIN_BORDER

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"
        data_start_row = 3
    else:
        data_start_row = 2

    for i, row in enumerate(df.itertuples(index=False), start=data_start_row):
        ws.append(list(row))
        alt = (i % 2 == 0)
        for col_idx, col_name in enumerate(headers, start=1):
            _style_body_cell(ws.cell(row=i, column=col_idx),
                             col_name=col_name, alt=alt)

    _autofit(ws)


def _summary_block(ws, label: str, grp: pd.DataFrame,
                   start_row: int, indent: int = 0) -> int:
    """
    Write a single summary block (label + metrics) starting at *start_row*.
    Returns the next free row number.
    """
    prefix = "  " * indent

    def _row(key, value):
        nonlocal start_row
        kc = ws.cell(row=start_row, column=2, value=f"{prefix}{key}")
        vc = ws.cell(row=start_row, column=3, value=value)
        kc.font = _BODY_FONT
        vc.font = _BODY_FONT
        kc.border = vc.border = _THIN_BORDER
        vc.alignment = Alignment(horizontal="center")
        start_row += 1

    # Section label
    lc = ws.cell(row=start_row, column=1, value=label)
    lc.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    lc.fill = _SUBHDR_FILL
    lc.border = _THIN_BORDER
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row,   end_column=3
    )
    start_row += 1

    _row("Pairs (n)",            len(grp))
    _row("Precision",            f"{grp['precision'].mean():.3f}")
    _row("Recall",               f"{grp['recall'].mean():.3f}")
    _row("F1",                   f"{grp['f1'].mean():.3f}")
    _row("Exact % (avg A+B)",    f"{_avg_exact(grp):.1f}%")
    _row("Seg mean (A) tokens",  _pooled_seg_mean(grp, "_a"))
    _row("Seg mean (B) tokens",  _pooled_seg_mean(grp, "_b"))
    start_row += 1   # blank spacer
    return start_row


def _write_summary_sheet(ws, df: pd.DataFrame, skipped: list[str]) -> None:
    """Sheet 2 — mirrors the terminal summary as a formatted table."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16

    # Title
    title_cell = ws.cell(row=1, column=1, value="Batch Evaluation Summary")
    title_cell.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = _HDR_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    # Metadata
    r = 3
    for label, value in [
        ("Document pairs evaluated", len(df)),
        ("Pairs skipped",            len(skipped)),
    ]:
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = _BOLD_FONT
        vc.font = _BODY_FONT
        r += 1

    if skipped:
        for s in skipped:
            sc = ws.cell(row=r, column=2, value=f"• {s}")
            sc.font = Font(name="Arial", size=9, color="C00000")
            r += 1

    r += 1

    if df.empty:
        ws.cell(row=r, column=1, value="No results to summarise.").font = _BODY_FONT
        return

    r = _summary_block(ws, "Overall (macro-averaged)", df, r)

    if df["variety"].nunique() > 1:
        for variety, grp in df.groupby("variety"):
            r = _summary_block(ws, f"Variety: {variety}", grp, r, indent=1)

    pair_col = df.apply(
        lambda row: " / ".join(sorted([row["annotator_a"], row["annotator_b"]])),
        axis=1,
    )
    if pair_col.nunique() > 1:
        for pair, grp in df.groupby(pair_col):
            r = _summary_block(ws, f"Pair: {pair}", grp, r, indent=1)


def _write_variety_sheet(ws, df: pd.DataFrame) -> None:
    """
    Sheet 3 — one row per language variety, adapt VARIETY_ORDER to control the type and order of varieties
    Includes all agreement metrics plus pooled segment-length statistics
    for both annotators, visually separated by a section divider row.
    """

    VARIETY_ORDER = ["E_HIST", "E_CURRENT", "DE_HIST", "DE_CURRENT"]

    agreement_metrics = [
        ("Pairs (n)",           lambda g: len(g)),
        ("Avg Spans A",         lambda g: round(g["spans_a"].mean(), 1)),
        ("Avg Spans B",         lambda g: round(g["spans_b"].mean(), 1)),
        ("Avg Exact Matches",   lambda g: round(g["exact_matches"].mean(), 1)),
        ("Exact % (avg A+B)",   lambda g: f"{_avg_exact(g):.1f}%"),
        ("Partial Overlaps %",  lambda g: f"{g['partial_overlaps_pct'].mean():.1f}%"),
        ("Only-A %",            lambda g: f"{g['only_a_pct'].mean():.1f}%"),
        ("Only-B %",            lambda g: f"{g['only_b_pct'].mean():.1f}%"),
        ("Precision",           lambda g: f"{g['precision'].mean():.3f}"),
        ("Recall",              lambda g: f"{g['recall'].mean():.3f}"),
        ("F1",                  lambda g: f"{g['f1'].mean():.3f}"),
    ]

    def _seg_metric(col, fmt=".1f", pct=False):
        def fn(g):
            if col not in g.columns:
                return "—"
            valid = g[col].dropna()
            if valid.empty:
                return "—"
            val = valid.mean()
            return f"{val:{fmt}}{'%' if pct else ''}"
        return fn

    seg_metrics_a = [
        ("Seg count A (avg)",       _seg_metric("seg_n_a", fmt=".1f")),
        ("Seg mean len A (tokens)", lambda g: _pooled_seg_mean(g, "_a")),
        ("Seg median len A",        _seg_metric("seg_median_a", fmt=".1f")),
        ("Seg min len A",           _seg_metric("seg_min_a", fmt=".1f")),
        ("Seg max len A",           _seg_metric("seg_max_a", fmt=".1f")),
        ("Seg std len A",           _seg_metric("seg_std_a", fmt=".2f")),
    ]
    seg_metrics_b = [
        ("Seg count B (avg)",       _seg_metric("seg_n_b", fmt=".1f")),
        ("Seg mean len B (tokens)", lambda g: _pooled_seg_mean(g, "_b")),
        ("Seg median len B",        _seg_metric("seg_median_b", fmt=".1f")),
        ("Seg min len B",           _seg_metric("seg_min_b", fmt=".1f")),
        ("Seg max len B",           _seg_metric("seg_max_b", fmt=".1f")),
        ("Seg std len B",           _seg_metric("seg_std_b", fmt=".2f")),
    ]

    # Determine which varieties are present
    present = [v for v in VARIETY_ORDER if v in df["variety"].values]
    for v in sorted(df["variety"].unique()):
        if v not in present:
            present.append(v)

    headers = ["Metric"] + present
    ws.append(headers)
    _style_header_row(ws, 1, ["Metric"] + [""] * len(present))  # plain blue for variety sheet
    # Re-colour header uniformly (variety sheet has no seg-len dedicated header row)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = _HDR_FILL
        cell.font   = _HDR_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    grouped = {v: df[df["variety"] == v] for v in present}

    def _write_metrics_block(metrics, row_offset, section_label=None,
                             header_fill=_HDR_FILL):
        if section_label:
            # Section divider row
            dc = ws.cell(row=row_offset, column=1, value=section_label)
            dc.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            dc.fill      = header_fill
            dc.border    = _THIN_BORDER
            dc.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(
                start_row=row_offset, start_column=1,
                end_row=row_offset,   end_column=len(headers)
            )
            ws.row_dimensions[row_offset].height = 18
            row_offset += 1

        for metric_label, fn in metrics:
            row_data = [metric_label] + [
                (fn(grouped[v]) if not grouped[v].empty else "—")
                for v in present
            ]
            ws.append(row_data)
            alt = (row_offset % 2 == 0)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_offset, column=col_idx)
                cell.border    = _THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_idx > 1 else "left",
                    vertical="center",
                )
                if col_idx == 1:
                    cell.font = _BOLD_FONT
                else:
                    cell.font = _BODY_FONT
                    cell.fill = _ALT_FILL if alt else PatternFill()
            row_offset += 1
        return row_offset

    current_row = 2
    current_row = _write_metrics_block(
        agreement_metrics, current_row,
        section_label="Agreement Metrics",
        header_fill=_SUBHDR_FILL,
    )
    current_row = _write_metrics_block(
        seg_metrics_a, current_row,
        section_label="Tokens per Segment — Annotator A",
        header_fill=_SEG_HDR_FILL,
    )
    current_row = _write_metrics_block(
        seg_metrics_b, current_row,
        section_label="Tokens per Segment — Annotator B",
        header_fill=_SEG_HDR_FILL,
    )

    _autofit(ws, min_width=14)


def write_xlsx(path: Path, df: pd.DataFrame, skipped: list[str]) -> None:
    """Write a three-sheet workbook to *path*."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results",    index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Summary",    index=False)
        pd.DataFrame().to_excel(writer, sheet_name="By Variety", index=False)

    wb = load_workbook(path)

    ws_results = wb["Results"]
    ws_results.delete_rows(1, ws_results.max_row)
    _write_results_sheet(ws_results, df)

    ws_summary = wb["Summary"]
    ws_summary.delete_rows(1, ws_summary.max_row)
    _write_summary_sheet(ws_summary, df, skipped)

    ws_variety = wb["By Variety"]
    ws_variety.delete_rows(1, ws_variety.max_row)
    _write_variety_sheet(ws_variety, df)

    wb.save(path)


# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Batch inter-annotator agreement evaluation"
    )
    parser.add_argument(
        "--root", type=Path, default=Path.cwd(),
        help="Directory containing the Data/ folder (default: cwd)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Discover pairs and print them without running evaluation",
    )
    parser.add_argument(
    "--include-dir", nargs="*", default=None,
    help="Only include paths containing these substrings (e.g. EN DE_HIST)"
    )
    parser.add_argument(
        "--exclude-dir", nargs="*", default=None,
        help="Exclude paths containing these substrings (e.g. WIP OLD)"
    )
    parser.add_argument(
        "--folder-whitelist", nargs="*", default=None,
        help="Only process annotator folders containing these tokens (e.g. DONE FINAL)"
    )
    parser.add_argument(
        "--folder-blacklist", nargs="*", default=None,
        help="Skip annotator folders containing these tokens (e.g. WIP TEMP)"
    )
    parser.add_argument(
        "--suffix", default="_DONE",
        help="Annotator folder suffix (default: _DONE)"
    )
    args = parser.parse_args()

    segmentation_bi = args.root / "Data" / "Segmentation_BI"
    out_stats       = args.root / "Results" / f"evaluation_results{timestr}.csv"
    out_xlsx        = args.root / "Results" / f"evaluation_results{timestr}.xlsx"
    out_nonmatches  = args.root / "Results" / f"non_matches{timestr}.csv"

    if not segmentation_bi.is_dir():
        sys.exit(f"Error: directory not found: {segmentation_bi}")

    print(f"Scanning {segmentation_bi}…")
    pairs = find_pairs(
    segmentation_bi,
    suffix=args.suffix,
    include_dir=args.include_dir,
    exclude_dir=args.exclude_dir,
    folder_whitelist=args.folder_whitelist,
    folder_blacklist=args.folder_blacklist,
)

    if not pairs:
        print("No matching annotator pairs found.")
        return

    print(f"Found {len(pairs)} pair(s) to evaluate.\n")

    if args.dry_run:
        for path_a, path_b, ann_a, ann_b, lang, variety in pairs:
            print(f"  [{lang} / {variety}]  {ann_a} vs {ann_b}:  {path_a.name}")
        print("\nDry run — nothing evaluated.")
        return

    stats_rows:    list[dict] = []
    nonmatch_rows: list[dict] = []
    skipped:       list[str]  = []

    for i, (path_a, path_b, ann_a, ann_b, lang, variety) in enumerate(pairs, 1):
        label = f"[{i}/{len(pairs)}] {ann_a} vs {ann_b}: {path_a.name}"
        print(label)

        try:
            result = evaluate_pair(path_a, path_b, verbose=False)
        except Exception as exc:
            print(f"Error: {exc}")
            skipped.append(f"{label} — {exc}")
            continue

        if result is None:
            mismatch_detail = find_text_mismatch(path_a, path_b)
            print(f"  Skipped — text mismatch: {mismatch_detail}")
            skipped.append(f"{label} — {mismatch_detail}")
            continue

        if result.ws_unified:
            print(f"Whitespace tokens unified")

        print(
            f"  spans: {result.spans_a} / {result.spans_b}  "
            f"exact: {result.exact} ({result.exact_pct:.1f}%)  "
            f"F1: {result.f1:.3f}"
        )

        stats_rows.append(
            result_to_stats_row(result, path_a, path_b, ann_a, ann_b,
                                lang, variety, segmentation_bi)
        )
        nonmatch_rows.extend(
            result_to_nonmatch_rows(result, path_a, ann_a, ann_b, lang, variety,
                                    segmentation_bi)
        )

    df_stats      = pd.DataFrame(stats_rows)
    df_nonmatches = pd.DataFrame(nonmatch_rows)

    if not df_stats.empty:
        df_stats.to_csv(out_stats, index=False, encoding="utf-8")
        print(f"\nStats written to       {out_stats}")

        write_xlsx(out_xlsx, df_stats, skipped)
        print(f"XLSX written to        {out_xlsx}")

    if not df_nonmatches.empty:
        df_nonmatches.to_csv(out_nonmatches, index=False, encoding="utf-8")
        print(f"Non-matches written to {out_nonmatches}")

    print_summary(df_stats, skipped)


if __name__ == "__main__":
    main()